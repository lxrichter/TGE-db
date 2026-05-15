import sqlite3
from pathlib import Path
from openpyxl import load_workbook

DB_PATH = Path("web/data/tge.db")
XLSX_PATH = Path("source-data/geothermal_companies.xlsx")

print(f"Using DB: {DB_PATH.resolve()}")
print(f"Using Excel: {XLSX_PATH.resolve()}")

if not DB_PATH.exists():
    raise FileNotFoundError(f"Database not found: {DB_PATH.resolve()}")

if not XLSX_PATH.exists():
    raise FileNotFoundError(f"Excel file not found: {XLSX_PATH.resolve()}")

SHEET_TO_TABLE = {
    "companies_master": "companies",
    "company_roles": "company_roles",
    "company_relationships": "company_relationships",
    "company_project_links": "company_project_links",
    "company_plant_links": "company_plant_links",
    "ref_company_type_primary": "ref_company_type_primary",
    "ref_company_type_secondary": "ref_company_type_secondary",
    "ref_company_roles": "ref_company_roles",
}

wb = load_workbook(XLSX_PATH, data_only=False)

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

try:
    cur.execute("PRAGMA foreign_keys = OFF;")
    conn.execute("BEGIN;")

    import_order = [
        "ref_company_type_primary",
        "ref_company_type_secondary",
        "ref_company_roles",
        "companies_master",
        "company_roles",
        "company_relationships",
        "company_project_links",
        "company_plant_links",
    ]

    for sheet_name in import_order:
        if sheet_name not in wb.sheetnames:
            print(f"[SKIP] Sheet not found: {sheet_name}")
            continue

        table_name = SHEET_TO_TABLE[sheet_name]
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"[SKIP] Empty sheet: {sheet_name}")
            continue

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = rows[1:]

        valid_indices = [i for i, h in enumerate(headers) if h]
        headers = [headers[i] for i in valid_indices]

        cur.execute(f"PRAGMA table_info({table_name});")
        db_columns = [row[1] for row in cur.fetchall()]

        common_columns = [h for h in headers if h in db_columns]

        if not common_columns:
            print(f"[SKIP] No matching columns for sheet {sheet_name} -> table {table_name}")
            continue

        cur.execute(f"DELETE FROM {table_name};")

        inserted = 0
        for raw_row in data_rows:
            if raw_row is None or all(v is None or str(v).strip() == "" for v in raw_row):
                continue

            row_dict = {}
            for original_idx in valid_indices:
                if original_idx < len(raw_row):
                    key = rows[0][original_idx]
                    if key is None:
                        continue
                    key = str(key).strip()
                    if key in common_columns:
                        value = raw_row[original_idx]
                        if isinstance(value, str):
                            value = value.strip()
                            if value == "":
                                value = None
                        row_dict[key] = value

            values = [row_dict.get(col, None) for col in common_columns]

            placeholders = ", ".join(["?"] * len(common_columns))
            col_sql = ", ".join(common_columns)

            cur.execute(
                f"INSERT INTO {table_name} ({col_sql}) VALUES ({placeholders})",
                values
            )
            inserted += 1

        print(f"[OK] Imported {inserted} rows from {sheet_name} -> {table_name}")

    conn.commit()
    print("[DONE] Import completed successfully.")

except Exception as e:
    conn.rollback()
    print("[ERROR] Import failed. Transaction rolled back.")
    print(e)
    raise

finally:
    cur.execute("PRAGMA foreign_keys = ON;")
    conn.close()

