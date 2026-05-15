from collections import Counter
from openpyxl import load_workbook
from pathlib import Path

XLSX_PATH = Path("source-data/geothermal_companies.xlsx")

wb = load_workbook(XLSX_PATH, data_only=False)
ws = wb["company_roles"]

rows = list(ws.iter_rows(values_only=True))
headers = [str(h).strip() if h is not None else "" for h in rows[0]]

try:
    idx = headers.index("company_role_id")
except ValueError:
    raise ValueError("Column 'company_role_id' not found in company_roles sheet")

ids = []
for row_num, row in enumerate(rows[1:], start=2):
    if row is None:
        continue
    value = row[idx] if idx < len(row) else None
    if value is None or str(value).strip() == "":
        continue
    ids.append((str(value).strip(), row_num))

counter = Counter(v for v, _ in ids)
dupes = {k: v for k, v in counter.items() if v > 1}

if not dupes:
    print("No duplicate company_role_id values found.")
else:
    print("Duplicate company_role_id values found:")
    for dup_id, count in sorted(dupes.items()):
        print(f"\n{dup_id} appears {count} times at rows:")
        for value, row_num in ids:
            if value == dup_id:
                print(f"  - Excel row {row_num}")
